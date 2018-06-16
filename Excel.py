# -*- coding: utf-8 -*-
__author__ = "1041656 張楷翊"

# Modules to Create Excel
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
# Module to Creating Filename
from datetime import datetime
# Modules we wrote
from Currency import currency
from StockClosingPrice import stock, stock2330
from WeatherForecast import weather


def size(sheet, times):
  for col in sheet.columns:
    max_length = 0
    column = col[0].column # Get the column name
    for cell in col:
      try: # Necessary to avoid error on empty cells
          if len(str(cell.value)) > max_length:
              max_length = len(cell.value)
      except:
          pass
    adjusted_width = (max_length + 2) * times
    sheet.column_dimensions[column].width = adjusted_width



if __name__ == "__main__":
  
  """ Datetime Now """
  t = datetime.now().strftime('%Y%m%d-%H%M')

  """ Get data form other Modules """
  stock = stock()
  stock2330 = stock2330()
  currency = currency()
  weather = weather()

  """ Set Titles """
  stockTitle = ['加權指數 (TAIEX)', '加權指數（含金融）', '加權指數（百分比）', '成交金額（億） (Changingover)']
  stock2330Title = ['股票代號 (Name)', '時間 (Time)', '成本', '買進 (Bid)', '賣出 (Ask)', '漲跌 (% Chg)', '張數', '昨收 (Prev Close)', '開盤 (Open)', '最高 (Highest)', '最低 (Lowest)']
  weatherTitle = ['日期 (Date)', '最高溫 (High)', '最低溫 (Low)', '天氣描述 (Description)']


  """ Create New Workbook and Worksheets """
  wb = Workbook()
  currencySheet = wb.create_sheet('Currency', 0)
  stockSheet = wb.create_sheet('Stock', 1)
  weatherSheet = wb.create_sheet('Weather', 2)

  """ Write Data to File """
  # Currency
  for r in dataframe_to_rows(currency, index=False, header=True):
    currencySheet.append(r)
  # TAIEX
  stockSheet.append(["大盤 (TAIEX)"])
  stockSheet.append(stockTitle)
  stockSheet.append(stock)
  stockSheet.append([])
  # Self-Selected Stock
  stockSheet.append(["自選股 (Self-Selected Stock)"])
  stockSheet.append(stock2330Title)
  stockSheet.append(stock2330)
  # Weather
  weatherSheet.append(weatherTitle)
  for i in range(0, 7):
    weatherSheet.append(weather[i])

  """ Styling Cells """
  stockSheet.merge_cells("A1:D1")
  stockSheet.merge_cells("A5:K5")

  # Setting Style (Title)
  titleStyle = NamedStyle(name='Title')
  titleStyle.font = Font(bold=True, vertAlign='baseline')
  titleStyle.alignment = Alignment(horizontal="center", vertical="center")

  # Apply Style (Title)
  wb.add_named_style(titleStyle)
  
  for cell in currencySheet['1']:
    cell.style = 'Title'
  stockSheet["A1"].style = 'Title'
  stockSheet["A5"].style = 'Title'
  for cell in weatherSheet['1']:
    cell.style = 'Title'

  # Auto Size Column Width
  size(currencySheet, 2.2)
  size(stockSheet, 1.8)
  size(weatherSheet, 1.5)

  # Remove Default WorkSheet when Creating Workbook
  wb.remove(wb["Sheet"])
  # Write out Excel File
  wb.save(t+".xlsx")